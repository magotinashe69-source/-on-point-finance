"""Excel export (openpyxl). Amounts are real numbers so the sheet can be summed."""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.money import cents_to_major

NAVY_HEX = "16264D"
MONEY_FORMAT = "#,##0.00"


def build_excel(start_day: date, end_day: date, entries, totals: dict, breakdown: dict) -> bytes:
    """Render the report to .xlsx bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Entries"

    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    navy_fill = PatternFill("solid", fgColor=NAVY_HEX)
    right = Alignment(horizontal="right")

    # Branding / range header.
    ws["A1"] = "On Point Educational Centre"
    ws["A1"].font = Font(bold=True, size=14, color=NAVY_HEX)
    ws["A2"] = "Quality Beyond Measure"
    ws["A3"] = f"Financial record: {start_day.isoformat()} to {end_day.isoformat()}"

    headers = ["Date", "Type", "Description", "Category", "Method", "Amount (MT)"]
    header_row = 5
    for col, text in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=text)
        c.font = white_bold
        c.fill = navy_fill

    row = header_row + 1
    for e in entries:
        ws.cell(row=row, column=1, value=e.date.isoformat())
        ws.cell(row=row, column=2, value=e.type.capitalize())
        ws.cell(row=row, column=3, value=e.description or "")
        ws.cell(row=row, column=4, value=e.category.name)
        ws.cell(row=row, column=5, value=e.payment_method)
        amount = ws.cell(row=row, column=6, value=cents_to_major(e.amount_cents))
        amount.number_format = MONEY_FORMAT
        row += 1

    # Totals block.
    row += 1
    for label, cents in (
        ("Income", totals["income_cents"]),
        ("Expenses", totals["expense_cents"]),
        ("Balance", totals["balance_cents"]),
    ):
        lbl = ws.cell(row=row, column=5, value=label)
        lbl.font = bold
        lbl.alignment = right
        val = ws.cell(row=row, column=6, value=cents_to_major(cents))
        val.number_format = MONEY_FORMAT
        val.font = bold
        row += 1

    widths = [12, 10, 34, 18, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
